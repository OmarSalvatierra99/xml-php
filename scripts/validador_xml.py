#!/usr/bin/env python3
"""
Validador de CFDI con el servicio del SAT
Verifica autenticidad y estatus fiscal de comprobantes digitales
"""

import os
import sys
import json
import pandas as pd
from datetime import datetime
from xml_utils import IssueTracker, load_xml_root, find_first, find_first_local, strip_namespace, get_attr, print_progress

# Namespaces
NAMESPACES_CFDI_40 = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital",
}

NAMESPACES_CFDI_33 = {
    "cfdi": "http://www.sat.gob.mx/cfd/3",
    "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital",
}


def extraer_datos_cfdi(filepath: str, tracker: IssueTracker) -> dict:
    """
    Extrae los datos necesarios para validación: UUID, RFCs, Total

    Args:
        filepath: Ruta al archivo XML
        tracker: IssueTracker para registrar problemas

    Returns:
        Dict con los datos extraídos o None si falla
    """
    root = load_xml_root(filepath, tracker)
    if root is None:
        return None

    filename = os.path.basename(filepath)

    # Intentar con CFDI 4.0
    comprobante = find_first(root, ".//cfdi:Comprobante", NAMESPACES_CFDI_40)
    namespaces = NAMESPACES_CFDI_40

    # Si no es 4.0, intentar con 3.3
    if comprobante is None:
        comprobante = find_first(root, ".//cfdi:Comprobante", NAMESPACES_CFDI_33)
        namespaces = NAMESPACES_CFDI_33

    # Si el comprobante es el nodo raíz o no se encontró por namespaces, buscar por nombre local
    if comprobante is None and strip_namespace(root.tag) == "Comprobante":
        comprobante = root
    if comprobante is None:
        comprobante = find_first_local(root, "Comprobante")

    if comprobante is None:
        tracker.error(f"'{filename}' no es un CFDI válido (no se encontró Comprobante)")
        return None

    # Extraer datos del Comprobante
    total = get_attr(comprobante, 'Total', '0.0')

    # Extraer Emisor
    emisor = find_first(root, ".//cfdi:Emisor", namespaces)
    if emisor is None:
        emisor = find_first_local(root, "Emisor")
    rfc_emisor = get_attr(emisor, 'Rfc', '') if emisor is not None else ''
    nombre_emisor = get_attr(emisor, 'Nombre', '') if emisor is not None else ''

    # Extraer Receptor
    receptor = find_first(root, ".//cfdi:Receptor", namespaces)
    if receptor is None:
        receptor = find_first_local(root, "Receptor")
    rfc_receptor = get_attr(receptor, 'Rfc', '') if receptor is not None else ''
    nombre_receptor = get_attr(receptor, 'Nombre', '') if receptor is not None else ''

    # Extraer UUID del TimbreFiscalDigital
    tfd = find_first(root, ".//tfd:TimbreFiscalDigital", namespaces)
    if tfd is None:
        tfd = find_first_local(root, "TimbreFiscalDigital")
    uuid = get_attr(tfd, 'UUID', '') if tfd is not None else ''

    if not uuid:
        tracker.error(f"'{filename}' no tiene UUID (TimbreFiscalDigital)")
        return None

    if not rfc_emisor or not rfc_receptor or not total:
        tracker.warn(f"'{filename}' tiene datos incompletos (RFC Emisor, Receptor o Total faltantes)")

    return {
        'archivo': filename,
        'uuid': uuid,
        'rfc_emisor': rfc_emisor,
        'nombre_emisor': nombre_emisor,
        'rfc_receptor': rfc_receptor,
        'nombre_receptor': nombre_receptor,
        'total': total,
        'estatus': '',
        'codigo_estatus': '',
        'es_cancelable': '',
        'estado_cancelacion': '',
        'fecha_validacion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }


def validar_con_sat(uuid: str, rfc_emisor: str, rfc_receptor: str, total: str, tracker: IssueTracker) -> dict:
    """
    Valida un CFDI con el servicio web del SAT

    Args:
        uuid: UUID del CFDI
        rfc_emisor: RFC del emisor
        rfc_receptor: RFC del receptor
        total: Monto total del CFDI
        tracker: IssueTracker para registrar problemas

    Returns:
        Dict con el resultado de la validación
    """
    try:
        # Importar zeep (SOAP client)
        from zeep import Client
        from zeep.transports import Transport
        from requests import Session
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry

    except ImportError:
        tracker.fatal("La biblioteca 'zeep' no está instalada. Ejecute: pip install zeep")
        return {
            'estatus': 'ERROR',
            'codigo_estatus': 'N/A',
            'es_cancelable': 'N/A',
            'estado_cancelacion': 'N/A'
        }

    try:
        # Configurar sesión con reintentos
        session = Session()
        retry = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
        adapter = HTTPAdapter(max_retries=retry)
        session.mount('http://', adapter)
        session.mount('https://', adapter)

        transport = Transport(session=session, timeout=30)

        # URL del servicio del SAT
        wsdl = 'https://consultaqr.facturaelectronica.sat.gob.mx/ConsultaCFDIService.svc?wsdl'

        # Crear cliente SOAP
        client = Client(wsdl=wsdl, transport=transport)

        # Preparar parámetros (expresión de consulta)
        # Formato: ?re=RFC_EMISOR&rr=RFC_RECEPTOR&tt=TOTAL&id=UUID
        expresion = f"?re={rfc_emisor}&rr={rfc_receptor}&tt={total}&id={uuid}"

        print_progress(f"  Consultando SAT para UUID: {uuid[:8]}...")

        # Llamar al servicio
        response = client.service.Consulta(expresion)

        # Parsear respuesta
        codigo_estatus = response.CodigoEstatus if hasattr(response, 'CodigoEstatus') else 'N/A'
        estado = response.Estado if hasattr(response, 'Estado') else 'N/A'
        es_cancelable = response.EsCancelable if hasattr(response, 'EsCancelable') else 'N/A'
        estado_cancelacion = response.EstatusCalificacion if hasattr(response, 'EstatusCalificacion') else 'N/A'

        # Mapear estado a texto comprensible
        if codigo_estatus == 'S - Comprobante obtenido satisfactoriamente.' or 'Vigente' in str(estado):
            estatus_texto = 'Vigente'
        elif codigo_estatus == 'N - 601: La consulta del comprobante resultó No encontrado.':
            estatus_texto = 'No encontrado'
        elif 'Cancelado' in str(estado):
            estatus_texto = 'Cancelado'
        else:
            estatus_texto = str(estado)

        return {
            'estatus': estatus_texto,
            'codigo_estatus': str(codigo_estatus),
            'es_cancelable': str(es_cancelable),
            'estado_cancelacion': str(estado_cancelacion)
        }

    except Exception as e:
        tracker.warn(f"Error al validar UUID {uuid[:8]}: {e}")
        return {
            'estatus': 'Error de conexión',
            'codigo_estatus': str(e),
            'es_cancelable': 'N/A',
            'estado_cancelacion': 'N/A'
        }


def validar_archivos(workdir: str, tracker: IssueTracker) -> str:
    """
    Valida todos los archivos XML en el directorio

    Args:
        workdir: Directorio con los XMLs a validar
        tracker: IssueTracker para registrar problemas

    Returns:
        Path del archivo Excel generado
    """
    # Obtener lista de archivos XML
    xml_files = [f for f in os.listdir(workdir) if f.lower().endswith('.xml')]

    if not xml_files:
        tracker.error("No se encontraron archivos XML en el directorio")
        return None

    print_progress(f"Validando {len(xml_files)} archivo(s) XML con el SAT...")

    resultados = []
    stats = {
        'vigente': 0,
        'cancelado': 0,
        'no_encontrado': 0,
        'error': 0
    }

    # Procesar cada archivo
    for filename in xml_files:
        filepath = os.path.join(workdir, filename)

        # Extraer datos del CFDI
        datos = extraer_datos_cfdi(filepath, tracker)

        if datos is None:
            # Error al procesar archivo
            resultados.append({
                'archivo': filename,
                'uuid': 'N/A',
                'rfc_emisor': 'N/A',
                'nombre_emisor': 'N/A',
                'rfc_receptor': 'N/A',
                'nombre_receptor': 'N/A',
                'total': 'N/A',
                'estatus': 'ERROR',
                'codigo_estatus': 'No se pudo leer el archivo',
                'es_cancelable': 'N/A',
                'estado_cancelacion': 'N/A',
                'fecha_validacion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            stats['error'] += 1
            continue

        # Validar con SAT
        validacion = validar_con_sat(
            datos['uuid'],
            datos['rfc_emisor'],
            datos['rfc_receptor'],
            datos['total'],
            tracker
        )

        # Actualizar datos con resultado de validación
        datos.update(validacion)

        # Actualizar estadísticas
        if datos['estatus'] == 'Vigente':
            stats['vigente'] += 1
        elif datos['estatus'] == 'Cancelado':
            stats['cancelado'] += 1
        elif datos['estatus'] == 'No encontrado':
            stats['no_encontrado'] += 1
        else:
            stats['error'] += 1

        resultados.append(datos)

        print_progress(f"  ✓ {filename} → {datos['estatus']}")

    # Crear DataFrame
    df = pd.DataFrame(resultados)

    # Ordenar columnas
    columnas_orden = [
        'archivo', 'estatus', 'uuid',
        'rfc_emisor', 'nombre_emisor',
        'rfc_receptor', 'nombre_receptor',
        'total', 'codigo_estatus',
        'es_cancelable', 'estado_cancelacion',
        'fecha_validacion'
    ]

    df = df[columnas_orden]

    # Generar archivo Excel
    excel_filename = 'Validacion_CFDI.xlsx'
    excel_path = os.path.join(workdir, excel_filename)

    print_progress(f"\nGenerando reporte Excel...")

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Validación')

            # Obtener el worksheet para aplicar formato
            workbook = writer.book
            worksheet = writer.sheets['Validación']

            # Aplicar formato condicional por estatus
            from openpyxl.styles import PatternFill, Font

            # Color coding
            vigente_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            cancelado_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            no_encontrado_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            error_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # Aplicar colores a las filas según estatus
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(columnas_orden)), start=2):
                estatus = worksheet.cell(row=idx, column=2).value

                if estatus == 'Vigente':
                    fill = vigente_fill
                elif estatus == 'Cancelado':
                    fill = cancelado_fill
                elif estatus == 'No encontrado':
                    fill = no_encontrado_fill
                else:
                    fill = error_fill

                for cell in row:
                    cell.fill = fill

            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        print_progress(f"✓ Reporte generado: {excel_filename}")
        print_progress(f"\nEstadísticas:")
        print_progress(f"  - Vigentes: {stats['vigente']}")
        print_progress(f"  - Cancelados: {stats['cancelado']}")
        print_progress(f"  - No encontrados: {stats['no_encontrado']}")
        print_progress(f"  - Errores: {stats['error']}")

        return {'excel_path': excel_path, 'stats': stats}

    except Exception as e:
        tracker.fatal(f"Error al crear Excel: {e}")
        return None


def main():
    if len(sys.argv) < 2:
        print("ERROR: Falta el directorio de trabajo", file=sys.stderr)
        sys.exit(2)

    workdir = sys.argv[1]

    if not os.path.isdir(workdir):
        print(f"ERROR: '{workdir}' no es un directorio válido", file=sys.stderr)
        sys.exit(2)

    tracker = IssueTracker()

    try:
        result = validar_archivos(workdir, tracker)

        # Reportar problemas
        tracker.report()

        # Output JSON con path y stats
        if result and 'excel_path' in result:
            output = {
                'path': result['excel_path'],
                'stats': result['stats']
            }
            print(json.dumps(output))
        else:
            print(json.dumps({'error': 'No se pudo generar el reporte'}))

        sys.exit(tracker.exit_code)

    except Exception as e:
        print(f"FATAL: Error inesperado: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
