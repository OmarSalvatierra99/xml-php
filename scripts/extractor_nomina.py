import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font
from typing import Dict, List, Optional, Set, Tuple

from xml_utils import (
    IssueTracker,
    find_all,
    find_all_local,
    find_first,
    find_first_local,
    get_attr,
    load_xml_root,
    print_progress,
    summarize_namespaces,
    to_float,
)

# Colores para celdas
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

NAMESPACES: Dict[str, str] = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "cfdi3": "http://www.sat.gob.mx/cfd/3",
    "nomina12": "http://www.sat.gob.mx/nomina12",
    "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital",
}


def _first_by_local_attr(root, local_name: str, must_have_any: Tuple[str, ...]) -> Optional[object]:
    candidates = find_all_local(root, local_name)
    for elem in candidates:
        if any(get_attr(elem, attr) is not None for attr in must_have_any):
            return elem
    return candidates[0] if candidates else None


def _nomina_elements(root, tag_name: str) -> List[object]:
    elems = find_all(root, f".//nomina12:{tag_name}", NAMESPACES)
    if not elems:
        elems = find_all_local(root, tag_name)
    return elems


def procesar_nomina_xml(directorio: str, tracker: IssueTracker) -> Optional[str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Perc_Deduc_Sub"
    ws.append(
        [
            "Archivo",
            "Tipo",
            "TipoPercepcion/Deduccion/Subsidio",
            "Clave",
            "Concepto",
            "ImporteGravado",
            "ImporteExento",
            "ImporteTotal",
        ]
    )

    percepciones_catalogo: Set[Tuple[str, str]] = set()
    deducciones_catalogo: Set[Tuple[str, str]] = set()
    subsidios_catalogo: Set[Tuple[str, str]] = set()

    xml_files = sorted(file for file in os.listdir(directorio) if file.lower().endswith(".xml"))
    if not xml_files:
        tracker.fatal(f"No se encontraron archivos XML en {directorio}")
        return None

    archivos_con_error: List[Tuple[str, str]] = []

    for filename in xml_files:
        ruta_archivo = os.path.join(directorio, filename)
        print_progress(f"Procesando nómina: {filename}")
        root = load_xml_root(ruta_archivo, tracker)
        if root is None:
            continue

        try:
            percepciones = _nomina_elements(root, "Percepcion")
            deducciones = _nomina_elements(root, "Deduccion")
            otros_pagos = _nomina_elements(root, "OtroPago")

            if not (percepciones or deducciones or otros_pagos):
                tracker.warn(
                    f"{filename}: No se detectaron nodos de nómina. Namespaces encontrados: {summarize_namespaces(root)}"
                )

            for percepcion in percepciones:
                tipo_percepcion = get_attr(percepcion, "TipoPercepcion") or ""
                clave = get_attr(percepcion, "Clave") or ""
                concepto = get_attr(percepcion, "Concepto") or ""
                importe_gravado = to_float(get_attr(percepcion, "ImporteGravado"), 0.0, tracker, "ImporteGravado")
                importe_exento = to_float(get_attr(percepcion, "ImporteExento"), 0.0, tracker, "ImporteExento")
                importe_total = importe_gravado + importe_exento
                ws.append(
                    [
                        filename,
                        "Percepción",
                        tipo_percepcion,
                        clave,
                        concepto,
                        importe_gravado,
                        importe_exento,
                        importe_total,
                    ]
                )
                ws.cell(row=ws.max_row, column=2).fill = green_fill
                if clave and concepto:
                    percepciones_catalogo.add((clave, concepto))

            for deduccion in deducciones:
                tipo_deduccion = get_attr(deduccion, "TipoDeduccion") or ""
                clave = get_attr(deduccion, "Clave") or ""
                concepto = get_attr(deduccion, "Concepto") or ""
                importe = to_float(get_attr(deduccion, "Importe"), 0.0, tracker, "Importe Deducción")
                ws.append([filename, "Deducción", tipo_deduccion, clave, concepto, "", "", importe])
                ws.cell(row=ws.max_row, column=2).fill = red_fill
                if clave and concepto:
                    deducciones_catalogo.add((clave, concepto))

            for otro_pago in otros_pagos:
                tipo_otro_pago = get_attr(otro_pago, "TipoOtroPago") or ""
                if tipo_otro_pago == "002":
                    clave = get_attr(otro_pago, "Clave") or ""
                    concepto = get_attr(otro_pago, "Concepto") or ""
                    importe = to_float(get_attr(otro_pago, "Importe"), 0.0, tracker, "Importe Subsidio")
                    ws.append([filename, "Subsidio", tipo_otro_pago, clave, concepto, "", "", importe])
                    ws.cell(row=ws.max_row, column=2).fill = blue_fill
                    if clave and concepto:
                        subsidios_catalogo.add((clave, concepto))

        except Exception as exc:
            archivos_con_error.append((filename, str(exc)))
            tracker.error(f"Error procesando {filename}: {exc}")

    catalog_ws = wb.create_sheet("Catalogo")
    catalog_ws.append(["Código", "Clave", "Concepto"])
    for clave, concepto in sorted(percepciones_catalogo):
        catalog_ws.append([f"P-{clave[:20]}", clave[:20], concepto[:50]])
    catalog_ws.append([])
    for clave, concepto in sorted(deducciones_catalogo):
        catalog_ws.append([f"D-{clave[:20]}", clave[:20], concepto[:50]])
    catalog_ws.append([])
    for clave, concepto in sorted(subsidios_catalogo):
        catalog_ws.append([f"S-{clave[:20]}", clave[:20], concepto[:50]])

    nomina_ws = wb.create_sheet("Nomina")
    nomina_headers = [
        "UUID",
        "Consecutivo",
        "Núm Empleado",
        "Nombre",
        "RFC",
        "CURP",
        "Puesto",
        "Departamento",
        "Tipo de Nomina",
        "Fecha Comprobante",
        "Num Días Pagados",
        "Fecha Inicial Pago",
        "Fecha Final Pago",
        "Fecha Pago",
        "Total Percepciones",
        "Total Deducciones",
        "Total Subsidios",
        "Total Neto",
    ]
    conceptos_headers: List[str] = []
    for clave, concepto in sorted(percepciones_catalogo):
        conceptos_headers.append(f"P-{clave[:15]}-{concepto[:20]}")
    for clave, concepto in sorted(deducciones_catalogo):
        conceptos_headers.append(f"D-{clave[:15]}-{concepto[:20]}")
    for clave, concepto in sorted(subsidios_catalogo):
        conceptos_headers.append(f"S-{clave[:15]}-{concepto[:20]}")

    nomina_ws.append(nomina_headers + conceptos_headers)
    for cell in nomina_ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    consecutivo = 1
    for filename in xml_files:
        ruta_archivo = os.path.join(directorio, filename)
        root = load_xml_root(ruta_archivo, tracker)
        if root is None:
            continue

        receptor_cfdi = find_first(root, ".//cfdi:Receptor", NAMESPACES)
        if receptor_cfdi is None:
            receptor_cfdi = find_first(root, ".//cfdi3:Receptor", NAMESPACES)
        if receptor_cfdi is None:
            receptor_cfdi = _first_by_local_attr(
                root, "Receptor", ("UsoCFDI", "RegimenFiscalReceptor", "DomicilioFiscalReceptor")
            )

        receptor_nomina = find_first(root, ".//nomina12:Receptor", NAMESPACES)
        if receptor_nomina is None:
            receptor_nomina = _first_by_local_attr(root, "Receptor", ("NumEmpleado", "Curp"))

        nomina = find_first(root, ".//nomina12:Nomina", NAMESPACES)
        if nomina is None:
            nomina = find_first_local(root, "Nomina")

        if receptor_cfdi is None or receptor_nomina is None or nomina is None:
            tracker.warn(
                f"Estructura de nómina incompleta en {filename}. "
                f"Receptor CFDI: {'OK' if receptor_cfdi is not None else 'No'}; "
                f"Receptor Nómina: {'OK' if receptor_nomina is not None else 'No'}; "
                f"Nomina: {'OK' if nomina is not None else 'No'}."
            )
            continue

        tfd = find_first(root, ".//tfd:TimbreFiscalDigital", NAMESPACES)
        if tfd is None:
            tfd = find_first_local(root, "TimbreFiscalDigital")
        uuid = get_attr(tfd, "UUID") or ""

        num_empleado = get_attr(receptor_nomina, "NumEmpleado") or ""
        nombre = get_attr(receptor_cfdi, "Nombre") or ""
        rfc = get_attr(receptor_cfdi, "Rfc") or ""
        curp = get_attr(receptor_nomina, "Curp") or ""
        puesto = get_attr(receptor_nomina, "Puesto") or ""
        departamento = get_attr(receptor_nomina, "Departamento") or ""
        tipo_nomina = get_attr(nomina, "TipoNomina") or ""
        fecha_comprobante = get_attr(root, "Fecha") or ""
        num_dias_pagados = get_attr(nomina, "NumDiasPagados") or ""
        fecha_inicial_pago = get_attr(nomina, "FechaInicialPago") or ""
        fecha_final_pago = get_attr(nomina, "FechaFinalPago") or ""
        fecha_pago = get_attr(nomina, "FechaPago") or ""

        total_percepciones = to_float(get_attr(nomina, "TotalPercepciones"), 0.0, tracker, "TotalPercepciones")
        total_deducciones = to_float(get_attr(nomina, "TotalDeducciones"), 0.0, tracker, "TotalDeducciones")
        total_subsidios = 0.0
        for otro_pago in _nomina_elements(root, "OtroPago"):
            if get_attr(otro_pago, "TipoOtroPago") == "002":
                total_subsidios += to_float(get_attr(otro_pago, "Importe"), 0.0, tracker, "Subsidios")
        total_neto = total_percepciones - total_deducciones + total_subsidios

        conceptos_valores: Dict[str, float] = {header: "" for header in conceptos_headers}
        for percepcion in _nomina_elements(root, "Percepcion"):
            clave = get_attr(percepcion, "Clave") or ""
            concepto_texto = get_attr(percepcion, "Concepto") or ""
            importe_total = to_float(get_attr(percepcion, "ImporteGravado"), 0.0, tracker, "ImporteGravado") + to_float(
                get_attr(percepcion, "ImporteExento"), 0.0, tracker, "ImporteExento"
            )
            header_key = f"P-{clave[:15]}-{concepto_texto[:20]}"
            if header_key in conceptos_valores and conceptos_valores[header_key] in ("", None):
                conceptos_valores[header_key] = importe_total

        for deduccion in _nomina_elements(root, "Deduccion"):
            clave = get_attr(deduccion, "Clave") or ""
            concepto_texto = get_attr(deduccion, "Concepto") or ""
            importe = to_float(get_attr(deduccion, "Importe"), 0.0, tracker, "Importe Deducción")
            header_key = f"D-{clave[:15]}-{concepto_texto[:20]}"
            if header_key in conceptos_valores and conceptos_valores[header_key] in ("", None):
                conceptos_valores[header_key] = importe

        for otro_pago in _nomina_elements(root, "OtroPago"):
            if get_attr(otro_pago, "TipoOtroPago") != "002":
                continue
            clave = get_attr(otro_pago, "Clave") or ""
            concepto_texto = get_attr(otro_pago, "Concepto") or ""
            importe = to_float(get_attr(otro_pago, "Importe"), 0.0, tracker, "Importe Subsidio")
            header_key = f"S-{clave[:15]}-{concepto_texto[:20]}"
            if header_key in conceptos_valores and conceptos_valores[header_key] in ("", None):
                conceptos_valores[header_key] = importe

        nomina_ws.append(
            [
                uuid,
                consecutivo,
                num_empleado,
                nombre,
                rfc,
                curp,
                puesto,
                departamento,
                tipo_nomina,
                fecha_comprobante,
                num_dias_pagados,
                fecha_inicial_pago,
                fecha_final_pago,
                fecha_pago,
                total_percepciones,
                total_deducciones,
                total_subsidios,
                total_neto,
            ]
            + [conceptos_valores[header] for header in conceptos_headers]
        )
        consecutivo += 1

    output_path = os.path.join(directorio, "Percepciones_Deducciones_Subsidios.xlsx")
    try:
        wb.save(output_path)
    except Exception as exc:
        tracker.fatal(f"No se pudo guardar el archivo Excel: {exc}")
        return None

    if archivos_con_error:
        tracker.error(f"{len(archivos_con_error)} archivo(s) con error durante el procesamiento.")

    return output_path


if __name__ == "__main__":
    tracker = IssueTracker()

    if len(sys.argv) > 1:
        directorio = sys.argv[1]
    else:
        directorio = os.path.dirname(os.path.abspath(__file__))

    excel_file = procesar_nomina_xml(directorio, tracker)

    tracker.report("Nómina")

    if excel_file and tracker.exit_code == 0:
        print(excel_file)

    sys.exit(tracker.exit_code)
